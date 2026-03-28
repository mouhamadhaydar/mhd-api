# -*- coding: utf-8 -*-
import re
from difflib import SequenceMatcher

try:
    from openpyxl import load_workbook
except Exception as e:
    raise Exception("openpyxl is not installed: " + str(e))


LOOKUP_COLUMNS = [
    ("Items_EN", "User_Query"),
    ("Inventory_EN", "User_Query"),
    ("Transactions_EN", "User_Query"),
    ("Arabic_English", "Arabic_Query"),
    ("Arabic_English", "English_Translation"),
]

SUITEQL_TEMPLATES = {
    "item_lookup": """
SELECT
    i.id,
    i.itemid,
    i.displayname,
    i.description,
    BUILTIN.DF(i.itemtype) AS item_type
FROM item i
WHERE UPPER(i.itemid) = UPPER('{item_code}')
""".strip(),

    "item_description": """
SELECT
    i.itemid,
    i.description
FROM item i
WHERE UPPER(i.itemid) = UPPER('{item_code}')
""".strip(),

    "list_items": """
SELECT
    i.id,
    i.itemid,
    i.displayname,
    i.description
FROM item i
WHERE i.isinactive = 'F'
ORDER BY i.itemid
""".strip(),

    "inventory_on_hand": """
SELECT
    i.itemid,
    BUILTIN.DF(ib.location) AS location,
    SUM(ib.quantityonhand) AS quantityonhand
FROM inventorybalance ib
JOIN item i ON i.id = ib.item
WHERE UPPER(i.itemid) = UPPER('{item_code}')
GROUP BY i.itemid, BUILTIN.DF(ib.location)
ORDER BY location
""".strip(),

    "inventory_available": """
SELECT
    i.itemid,
    BUILTIN.DF(ib.location) AS location,
    SUM(ib.quantityavailable) AS quantityavailable
FROM inventorybalance ib
JOIN item i ON i.id = ib.item
WHERE UPPER(i.itemid) = UPPER('{item_code}')
GROUP BY i.itemid, BUILTIN.DF(ib.location)
ORDER BY location
""".strip(),

    "inventory_committed": """
SELECT
    i.itemid,
    BUILTIN.DF(ib.location) AS location,
    SUM(ib.quantityonhand - ib.quantityavailable) AS committed_quantity
FROM inventorybalance ib
JOIN item i ON i.id = ib.item
WHERE UPPER(i.itemid) = UPPER('{item_code}')
GROUP BY i.itemid, BUILTIN.DF(ib.location)
ORDER BY location
""".strip(),

    "stock_by_bin": """
SELECT
    i.itemid,
    BUILTIN.DF(ib.location) AS location,
    b.binnumber,
    SUM(ib.quantityonhand) AS quantityonhand
FROM inventorybalance ib
JOIN item i ON i.id = ib.item
LEFT JOIN bin b ON b.id = ib.binnumber
WHERE UPPER(i.itemid) = UPPER('{item_code}')
GROUP BY i.itemid, BUILTIN.DF(ib.location), b.binnumber
ORDER BY location, b.binnumber
""".strip(),

    "transaction_sales_order": """
SELECT
    t.tranid,
    t.trandate,
    BUILTIN.DF(t.entity) AS customer,
    t.foreigntotal,
    BUILTIN.DF(t.status) AS status
FROM transaction t
WHERE t.type = 'SalesOrd'
ORDER BY t.trandate DESC
""".strip(),

    "transaction_purchase_order": """
SELECT
    t.tranid,
    t.trandate,
    BUILTIN.DF(t.entity) AS vendor,
    t.foreigntotal,
    BUILTIN.DF(t.status) AS status
FROM transaction t
WHERE t.type = 'PurchOrd'
ORDER BY t.trandate DESC
""".strip(),
}

INTENT_HINTS = {
    "stock_by_location": "inventory_on_hand",
    "low_stock": "inventory_on_hand",
    "out_of_stock": "inventory_on_hand",
    "show_items": "list_items",
    "sales_orders": "transaction_sales_order",
    "purchase_orders": "transaction_purchase_order",
}


def safe_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_text(text):
    text = safe_text(text).lower()
    text = re.sub(r"[^\w\s\-/\.]", " ", text, flags=re.UNICODE)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def similarity(a, b):
    return SequenceMatcher(None, normalize_text(a), normalize_text(b)).ratio()


def extract_item_code(text):
    text = safe_text(text)

    patterns = [
        r"\bitem\s+([A-Za-z0-9._\-/]+)\b",
        r"\bsku\s+([A-Za-z0-9._\-/]+)\b",
        r"\bcode\s+([A-Za-z0-9._\-/]+)\b",
        r"\bصنف\s+([A-Za-z0-9._\-/]+)\b",
        r"\b([A-Z]{2,}[A-Z0-9._\-/]*)\b",
        r"\b([A-Za-z]{2,}\d{2,}[A-Za-z0-9._\-/]*)\b"
    ]

    for pattern in patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()

    return ""


def header_map_from_sheet(sheet):
    mapping = {}
    for col_idx, cell in enumerate(sheet[1], start=1):
        header = safe_text(cell.value)
        if header:
            mapping[header] = col_idx
    return mapping


def get_cell(sheet, row_idx, col_idx):
    if not col_idx:
        return ""
    return safe_text(sheet.cell(row=row_idx, column=col_idx).value)


def load_query_index(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    rows = []

    for sheet_name, query_col in LOOKUP_COLUMNS:
        if sheet_name not in wb.sheetnames:
            continue

        sheet = wb[sheet_name]
        headers = header_map_from_sheet(sheet)

        if query_col not in headers:
            continue

        query_col_idx = headers.get(query_col)
        normalized_intent_idx = headers.get("Normalized_Intent")
        domain_idx = headers.get("Domain")
        query_type_idx = headers.get("Query_Type")
        key_entities_idx = headers.get("Key_Entities")
        suggested_tables_idx = headers.get("Suggested_Tables")
        notes_idx = headers.get("Notes")

        for row_idx in range(2, sheet.max_row + 1):
            query_text = get_cell(sheet, row_idx, query_col_idx)
            if not query_text:
                continue

            rows.append({
                "sheet": sheet_name,
                "row_number": row_idx,
                "query_column": query_col,
                "query_text": query_text,
                "normalized_intent": get_cell(sheet, row_idx, normalized_intent_idx),
                "domain": get_cell(sheet, row_idx, domain_idx),
                "query_type": get_cell(sheet, row_idx, query_type_idx),
                "key_entities": get_cell(sheet, row_idx, key_entities_idx),
                "suggested_tables": get_cell(sheet, row_idx, suggested_tables_idx),
                "notes": get_cell(sheet, row_idx, notes_idx),
            })

    return rows


def find_best_match(user_query, query_index):
    best = None
    best_score = -1.0

    for row in query_index:
        score = similarity(user_query, row["query_text"])
        if score > best_score:
            best_score = score
            best = row

    if not best:
        return None

    result = dict(best)
    result["match_score"] = round(best_score, 4)
    return result


def pick_template(intent):
    intent = safe_text(intent)
    if intent in SUITEQL_TEMPLATES:
        return intent
    if intent in INTENT_HINTS:
        return INTENT_HINTS[intent]
    return ""


def extract_item_from_entities(key_entities):
    text = safe_text(key_entities)
    m = re.search(r"item_code=([A-Za-z0-9._\-/]+)", text, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return ""


def generate_suiteql(user_query, matched):
    if not matched:
        return ""

    intent = matched.get("normalized_intent", "")
    template_key = pick_template(intent)

    if not template_key:
        return ""

    template = SUITEQL_TEMPLATES[template_key]

    item_code = extract_item_code(user_query)
    if not item_code:
        item_code = extract_item_from_entities(matched.get("key_entities", ""))

    if not item_code:
        item_code = "IPC0025"

    return template.format(item_code=item_code).strip()


def fetch_suiteql(user_query, excel_path):
    query_index = load_query_index(excel_path)
    matched = find_best_match(user_query, query_index)

    if not matched:
        return {
            "success": False,
            "error": "No matching query found",
            "suiteql": ""
        }

    suiteql = generate_suiteql(user_query, matched)

    return {
        "success": True,
        "user_query": user_query,
        "matched_query": matched.get("query_text", ""),
        "match_score": matched.get("match_score", 0),
        "sheet": matched.get("sheet", ""),
        "row_number": matched.get("row_number", ""),
        "domain": matched.get("domain", ""),
        "query_type": matched.get("query_type", ""),
        "normalized_intent": matched.get("normalized_intent", ""),
        "key_entities": matched.get("key_entities", ""),
        "suggested_tables": matched.get("suggested_tables", ""),
        "notes": matched.get("notes", ""),
        "suiteql": suiteql
    }
